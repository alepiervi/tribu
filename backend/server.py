from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional, Dict, Any, Union
from datetime import datetime, timedelta, timezone
from passlib.hash import bcrypt
import jwt
import os
import uuid
from dotenv import load_dotenv
from pathlib import Path
import shutil
from enum import Enum

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'your-secret-key')
JWT_ALGORITHM = 'HS256'
security = HTTPBearer()

# Create the main app
app = FastAPI(title="Travel Agency API")
api_router = APIRouter(prefix="/api")

# Enums
class UserRole(str, Enum):
    ADMIN = "admin"
    AGENT = "agent"
    CLIENT = "client"

class TripType(str, Enum):
    CRUISE = "cruise"
    RESORT = "resort"
    TOUR = "tour"
    CUSTOM = "custom"

class TripStatus(str, Enum):
    DRAFT = "draft"
    ACTIVE = "active"
    CONFIRMED = "confirmed"
    COMPLETED = "completed"
    CANCELLED = "cancelled"

class ItineraryType(str, Enum):
    PORT_DAY = "port_day"
    SEA_DAY = "sea_day"
    RESORT_DAY = "resort_day"
    TOUR_DAY = "tour_day"
    FREE_DAY = "free_day"

class POICategory(str, Enum):
    RESTAURANT = "restaurant"
    ATTRACTION = "attraction"
    HOTEL = "hotel"
    ACTIVITY = "activity"
    TRANSPORT = "transport"
    SHIP_FACILITY = "ship_facility"

class PhotoCategory(str, Enum):
    DESTINATION = "destination"
    SHIP_CABIN = "ship_cabin"
    SHIP_FACILITIES = "ship_facilities"
    DINING = "dining"
    ACTIVITIES = "activities"
    EXCURSION = "excursion"
    RESORT_ROOM = "resort_room"
    RESORT_BEACH = "resort_beach"
    RESORT_POOL = "resort_pool"
    TOUR_ATTRACTIONS = "tour_attractions"
    TOUR_HOTELS = "tour_hotels"

# Models
class User(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    email: EmailStr
    first_name: str
    last_name: str
    role: UserRole
    blocked: bool = False
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class UserCreate(BaseModel):
    email: EmailStr
    password: str
    first_name: str
    last_name: str
    role: UserRole = UserRole.CLIENT

class UserUpdate(BaseModel):
    first_name: Optional[str] = None
    last_name: Optional[str] = None
    email: Optional[EmailStr] = None
    blocked: Optional[bool] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class Trip(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    destination: str
    description: str
    start_date: datetime
    end_date: datetime
    client_id: str
    agent_id: str
    status: TripStatus = TripStatus.DRAFT
    trip_type: TripType
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TripCreate(BaseModel):
    title: str
    destination: str
    description: str
    start_date: datetime
    end_date: datetime
    client_id: str
    trip_type: TripType

class TripUpdate(BaseModel):
    title: Optional[str] = None
    destination: Optional[str] = None
    description: Optional[str] = None
    start_date: Optional[datetime] = None
    end_date: Optional[datetime] = None
    client_id: Optional[str] = None
    trip_type: Optional[TripType] = None
    status: Optional[TripStatus] = None

class Itinerary(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    trip_id: str
    day_number: int
    date: datetime
    title: str
    description: str
    itinerary_type: ItineraryType
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ItineraryCreate(BaseModel):
    trip_id: str
    day_number: int
    date: datetime
    title: str
    description: str
    itinerary_type: ItineraryType

class CruiseInfo(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    trip_id: str
    ship_name: str
    cabin_number: str
    departure_time: datetime
    return_time: datetime
    ship_facilities: List[str] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CruiseInfoCreate(BaseModel):
    trip_id: str
    ship_name: str
    cabin_number: str
    departure_time: datetime
    return_time: datetime
    ship_facilities: List[str] = []

class PortSchedule(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    trip_id: str
    itinerary_id: str
    port_name: str
    arrival_time: datetime
    departure_time: datetime
    all_aboard_time: datetime
    transport_info: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PortScheduleCreate(BaseModel):
    trip_id: str
    itinerary_id: str
    port_name: str
    arrival_time: datetime
    departure_time: datetime
    all_aboard_time: datetime
    transport_info: str = ""

class POI(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    category: POICategory
    address: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    description: str = ""
    phone: str = ""
    website: str = ""
    price_range: str = ""
    image_urls: List[str] = []
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class POICreate(BaseModel):
    name: str
    category: POICategory
    address: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    description: str = ""
    phone: str = ""
    website: str = ""
    price_range: str = ""
    image_urls: List[str] = []

class ItineraryPOI(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    itinerary_id: str
    poi_id: str
    order_number: int
    visit_time: datetime
    duration_minutes: int
    transport_type: str = ""
    transport_duration: int = 0
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ItineraryPOICreate(BaseModel):
    itinerary_id: str
    poi_id: str
    order_number: int
    visit_time: datetime
    duration_minutes: int
    transport_type: str = ""
    transport_duration: int = 0
    notes: str = ""

class ClientPhoto(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    trip_id: str
    client_id: str
    url: str
    caption: str = ""
    photo_category: PhotoCategory
    uploaded_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ShipActivity(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    cruise_info_id: str
    day_date: datetime
    activity_name: str
    activity_time: datetime
    location: str
    description: str = ""
    image_url: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ShipActivityCreate(BaseModel):
    cruise_info_id: str
    day_date: datetime
    activity_name: str
    activity_time: datetime
    location: str
    description: str = ""
    image_url: str = ""

class ClientNote(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    trip_id: str
    client_id: str
    day_number: int
    note_text: str
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ClientNoteCreate(BaseModel):
    trip_id: str
    day_number: int
    note_text: str

# Administrative/Financial Models
class TripAdmin(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    trip_id: str
    practice_number: str  # Numero scheda pratica
    booking_number: str  # Numero prenotazione
    gross_amount: float  # Importo lordo saldato
    net_amount: float  # Importo Netto
    discount: float  # Sconto
    gross_commission: float  # Commissione lorda (calculated)
    supplier_commission: float  # Commissione fornitore (calculated 4%)
    agent_commission: float  # Commissione Agente (calculated)
    practice_confirm_date: datetime  # Data conferma pratica
    client_departure_date: datetime  # Data partenza Cliente
    confirmation_deposit: float  # Acconto versato per conferma
    balance_due: float  # Saldo da versare (calculated)
    status: str = "draft"  # draft, confirmed, paid, cancelled
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TripAdminCreate(BaseModel):
    trip_id: str
    practice_number: str
    booking_number: str
    gross_amount: float
    net_amount: float
    discount: float = 0.0
    practice_confirm_date: datetime
    client_departure_date: datetime
    confirmation_deposit: float = 0.0

class TripAdminUpdate(BaseModel):
    practice_number: Optional[str] = None
    booking_number: Optional[str] = None
    gross_amount: Optional[float] = None
    net_amount: Optional[float] = None
    discount: Optional[float] = None
    practice_confirm_date: Optional[datetime] = None
    client_departure_date: Optional[datetime] = None
    confirmation_deposit: Optional[float] = None
    status: Optional[str] = None

class PaymentInstallment(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    trip_admin_id: str
    amount: float
    payment_date: datetime
    payment_type: str = "installment"  # installment, balance, deposit
    notes: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class PaymentInstallmentCreate(BaseModel):
    trip_admin_id: str
    amount: float
    payment_date: datetime
    payment_type: str = "installment"
    notes: str = ""

# Trip Details Models
class CruiseDetails(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    trip_id: str
    ship_name: str = ""
    boarding_port: str = ""
    cabin_number: str = ""
    package_type: str = ""
    insurance_type: str = ""
    restaurant: str = ""
    dinner_time: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CruiseDetailsUpdate(BaseModel):
    ship_name: Optional[str] = None
    boarding_port: Optional[str] = None
    cabin_number: Optional[str] = None
    package_type: Optional[str] = None
    insurance_type: Optional[str] = None
    restaurant: Optional[str] = None
    dinner_time: Optional[str] = None

class ResortDetails(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    trip_id: str
    resort_name: str = ""
    room_type: str = ""
    meal_plan: str = ""
    package_formula: str = ""
    insurance_type: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class ResortDetailsUpdate(BaseModel):
    resort_name: Optional[str] = None
    room_type: Optional[str] = None
    meal_plan: Optional[str] = None
    package_formula: Optional[str] = None
    insurance_type: Optional[str] = None

class TourDetails(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    trip_id: str
    general_info: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class TourDetailsUpdate(BaseModel):
    general_info: Optional[str] = None

class CustomTripDetails(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    trip_id: str
    custom_details: str = ""
    created_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))
    updated_at: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class CustomTripDetailsUpdate(BaseModel):
    custom_details: Optional[str] = None

# Utility functions
def create_token(user_data: dict) -> str:
    payload = {
        "user_id": user_data["id"],
        "email": user_data["email"],
        "role": user_data["role"],
        "exp": datetime.now(timezone.utc) + timedelta(days=7)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)) -> dict:
    try:
        payload = jwt.decode(credentials.credentials, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user = await db.users.find_one({"id": payload["user_id"]})
        if not user:
            raise HTTPException(status_code=401, detail="User not found")
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

def prepare_for_mongo(data):
    """Convert datetime objects to ISO strings for MongoDB storage"""
    if isinstance(data, dict):
        for key, value in data.items():
            if isinstance(value, datetime):
                data[key] = value.isoformat()
            elif isinstance(value, dict):
                data[key] = prepare_for_mongo(value)
            elif isinstance(value, list):
                data[key] = [prepare_for_mongo(item) if isinstance(item, dict) else item for item in value]
    return data

def calculate_trip_admin_fields(trip_admin_data: dict, installments: List[dict] = None) -> dict:
    """Calculate derived fields for trip administration"""
    gross_amount = trip_admin_data.get('gross_amount', 0)
    net_amount = trip_admin_data.get('net_amount', 0)
    discount = trip_admin_data.get('discount', 0)
    confirmation_deposit = trip_admin_data.get('confirmation_deposit', 0)
    
    # Calculate commissions
    gross_commission = gross_amount - discount - net_amount
    supplier_commission = gross_amount * 0.04  # 4% of gross
    agent_commission = gross_commission - supplier_commission
    
    # Calculate balance due
    total_paid = confirmation_deposit
    if installments:
        total_paid += sum(inst.get('amount', 0) for inst in installments)
    balance_due = gross_amount - total_paid
    
    return {
        **trip_admin_data,
        'gross_commission': gross_commission,
        'supplier_commission': supplier_commission,
        'agent_commission': agent_commission,
        'balance_due': balance_due
    }

def parse_from_mongo(item):
    """Parse datetime strings from MongoDB back to datetime objects and remove ObjectIds"""
    if isinstance(item, dict):
        # Remove MongoDB ObjectId field
        if '_id' in item:
            del item['_id']
        
        for key, value in item.items():
            if isinstance(value, str) and 'T' in value and (value.endswith('Z') or '+' in value):
                try:
                    item[key] = datetime.fromisoformat(value.replace('Z', '+00:00'))
                except:
                    pass
            elif isinstance(value, dict):
                item[key] = parse_from_mongo(value)
            elif isinstance(value, list):
                item[key] = [parse_from_mongo(subitem) if isinstance(subitem, dict) else subitem for subitem in value]
    return item

# Authentication endpoints
@api_router.post("/auth/register")
async def register(user_data: UserCreate):
    # Check if user exists
    existing_user = await db.users.find_one({"email": user_data.email})
    if existing_user:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Hash password
    hashed_password = bcrypt.hash(user_data.password)
    
    # Create user
    user = User(**user_data.dict(exclude={"password"}))
    user_dict = prepare_for_mongo(user.dict())
    user_dict["hashed_password"] = hashed_password
    
    await db.users.insert_one(user_dict)
    
    # Create token
    token = create_token(user_dict)
    
    return {"user": user, "token": token}

@api_router.post("/auth/login")
async def login(login_data: UserLogin):
    user = await db.users.find_one({"email": login_data.email})
    if not user or not bcrypt.verify(login_data.password, user["hashed_password"]):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    # Check if user is blocked
    if user.get("blocked", False):
        raise HTTPException(status_code=403, detail="Account blocked. Contact administrator.")
    
    token = create_token(user)
    user_response = User(**parse_from_mongo(user))
    
    return {"user": user_response, "token": token}

@api_router.get("/auth/me")
async def get_current_user_info(current_user: dict = Depends(get_current_user)):
    return User(**parse_from_mongo(current_user))

# Trip endpoints
@api_router.get("/trips", response_model=List[Trip])
async def get_trips(current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "admin":
        trips = await db.trips.find().to_list(1000)
    elif current_user["role"] == "agent":
        trips = await db.trips.find({"agent_id": current_user["id"]}).to_list(1000)
    else:  # client
        trips = await db.trips.find({"client_id": current_user["id"]}).to_list(1000)
    
    return [Trip(**parse_from_mongo(trip)) for trip in trips]

@api_router.get("/trips/with-details")
async def get_trips_with_details(current_user: dict = Depends(get_current_user)):
    """Get trips with agent and client details"""
    if current_user["role"] == "admin":
        trips = await db.trips.find().to_list(1000)
    elif current_user["role"] == "agent":
        trips = await db.trips.find({"agent_id": current_user["id"]}).to_list(1000)
    else:  # client
        trips = await db.trips.find({"client_id": current_user["id"]}).to_list(1000)
    
    # Get all unique user IDs
    agent_ids = list(set(trip["agent_id"] for trip in trips))
    client_ids = list(set(trip["client_id"] for trip in trips))
    
    # Fetch agents and clients
    agents = {}
    if agent_ids:
        agent_list = await db.users.find({"id": {"$in": agent_ids}}).to_list(1000)
        agents = {agent["id"]: {
            "id": agent["id"],
            "first_name": agent["first_name"],
            "last_name": agent["last_name"],
            "email": agent["email"]
        } for agent in agent_list}
    
    clients = {}
    if client_ids:
        client_list = await db.users.find({"id": {"$in": client_ids}}).to_list(1000)
        clients = {client["id"]: {
            "id": client["id"],
            "first_name": client["first_name"],
            "last_name": client["last_name"],
            "email": client["email"]
        } for client in client_list}
    
    # Combine trip data with user info
    trips_with_details = []
    for trip in trips:
        trip_data = Trip(**parse_from_mongo(trip))
        agent_info = agents.get(trip["agent_id"])
        client_info = clients.get(trip["client_id"])
        
        trips_with_details.append({
            "trip": trip_data,
            "agent": agent_info,
            "client": client_info
        })
    
    return trips_with_details

@api_router.post("/trips", response_model=Trip)
async def create_trip(trip_data: TripCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized to create trips")
    
    trip = Trip(**trip_data.dict(), agent_id=current_user["id"])
    trip_dict = prepare_for_mongo(trip.dict())
    
    await db.trips.insert_one(trip_dict)
    return trip

@api_router.get("/trips/{trip_id}", response_model=Trip)
async def get_trip(trip_id: str, current_user: dict = Depends(get_current_user)):
    trip = await db.trips.find_one({"id": trip_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Check permissions
    if (current_user["role"] == "client" and trip["client_id"] != current_user["id"]) or \
       (current_user["role"] == "agent" and trip["agent_id"] != current_user["id"]):
        raise HTTPException(status_code=403, detail="Not authorized to view this trip")
    
    return Trip(**parse_from_mongo(trip))

@api_router.get("/trips/{trip_id}/full", response_model=Dict[str, Any])
async def get_trip_with_details(trip_id: str, current_user: dict = Depends(get_current_user)):
    """Get trip with agent and client details"""
    trip = await db.trips.find_one({"id": trip_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Check permissions
    if (current_user["role"] == "client" and trip["client_id"] != current_user["id"]) or \
       (current_user["role"] == "agent" and trip["agent_id"] != current_user["id"]):
        raise HTTPException(status_code=403, detail="Not authorized to view this trip")
    
    # Get agent details
    agent = await db.users.find_one({"id": trip["agent_id"]})
    agent_info = None
    if agent:
        agent_info = {
            "id": agent["id"],
            "first_name": agent["first_name"],
            "last_name": agent["last_name"],
            "email": agent["email"]
        }
    
    # Get client details
    client = await db.users.find_one({"id": trip["client_id"]})
    client_info = None
    if client:
        client_info = {
            "id": client["id"],
            "first_name": client["first_name"],
            "last_name": client["last_name"],
            "email": client["email"]
        }
    
    return {
        "trip": Trip(**parse_from_mongo(trip)),
        "agent": agent_info,
        "client": client_info
    }

@api_router.put("/trips/{trip_id}", response_model=Trip)
async def update_trip(trip_id: str, trip_data: TripUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized to update trips")
    
    trip = await db.trips.find_one({"id": trip_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Check if agent is updating their own trip
    if current_user["role"] == "agent" and trip["agent_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Agents can only update their own trips")
    
    # Prepare update data, excluding None values
    update_data = {}
    for field, value in trip_data.dict(exclude_unset=True).items():
        if value is not None:
            if isinstance(value, datetime):
                update_data[field] = value.isoformat()
            else:
                update_data[field] = value
    
    if update_data:
        await db.trips.update_one({"id": trip_id}, {"$set": update_data})
    
    updated_trip = await db.trips.find_one({"id": trip_id})
    return Trip(**parse_from_mongo(updated_trip))

@api_router.delete("/trips/{trip_id}")
async def delete_trip(trip_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized to delete trips")
    
    # Check if trip exists
    trip = await db.trips.find_one({"id": trip_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # If agent, check if they own this trip
    if current_user["role"] == "agent" and trip["agent_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized to delete this trip")
    
    # Cascading delete: remove all related data
    print(f"🗑️ Deleting trip {trip_id} and all related data...")
    
    # Delete financial/admin data
    trip_admin_result = await db.trip_admin.delete_many({"trip_id": trip_id})
    print(f"   📊 Deleted {trip_admin_result.deleted_count} financial records")
    
    # Delete itineraries
    itinerary_result = await db.itineraries.delete_many({"trip_id": trip_id})
    print(f"   📅 Deleted {itinerary_result.deleted_count} itinerary records")
    
    # Delete cruise info
    cruise_result = await db.cruise_info.delete_many({"trip_id": trip_id})
    print(f"   🚢 Deleted {cruise_result.deleted_count} cruise info records")
    
    # Delete client notes
    notes_result = await db.client_notes.delete_many({"trip_id": trip_id})
    print(f"   📝 Deleted {notes_result.deleted_count} client notes")
    
    # Delete client photos
    photos_result = await db.client_photos.delete_many({"trip_id": trip_id})
    print(f"   📷 Deleted {photos_result.deleted_count} client photos")
    
    # Finally delete the trip itself
    result = await db.trips.delete_one({"id": trip_id})
    print(f"   ✈️ Deleted trip: {result.deleted_count}")
    
    return {
        "message": "Trip and all related data deleted successfully",
        "deleted_counts": {
            "trip": result.deleted_count,
            "financial_records": trip_admin_result.deleted_count,
            "itineraries": itinerary_result.deleted_count,
            "cruise_info": cruise_result.deleted_count,
            "client_notes": notes_result.deleted_count,
            "client_photos": photos_result.deleted_count
        }
    }

# Itinerary endpoints
@api_router.get("/trips/{trip_id}/itineraries", response_model=List[Itinerary])
async def get_itineraries(trip_id: str, current_user: dict = Depends(get_current_user)):
    itineraries = await db.itineraries.find({"trip_id": trip_id}).to_list(1000)
    return [Itinerary(**parse_from_mongo(itinerary)) for itinerary in itineraries]

@api_router.post("/itineraries", response_model=Itinerary)
async def create_itinerary(itinerary_data: ItineraryCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized to create itineraries")
    
    itinerary = Itinerary(**itinerary_data.dict())
    itinerary_dict = prepare_for_mongo(itinerary.dict())
    
    await db.itineraries.insert_one(itinerary_dict)
    return itinerary

@api_router.put("/itineraries/{itinerary_id}", response_model=Itinerary)
async def update_itinerary(itinerary_id: str, itinerary_data: ItineraryCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized to update itineraries")
    
    update_data = prepare_for_mongo(itinerary_data.dict())
    await db.itineraries.update_one({"id": itinerary_id}, {"$set": update_data})
    
    updated_itinerary = await db.itineraries.find_one({"id": itinerary_id})
    if not updated_itinerary:
        raise HTTPException(status_code=404, detail="Itinerary not found")
    
    return Itinerary(**parse_from_mongo(updated_itinerary))

@api_router.delete("/itineraries/{itinerary_id}")
async def delete_itinerary(itinerary_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized to delete itineraries")
    
    result = await db.itineraries.delete_one({"id": itinerary_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Itinerary not found")
    
    return {"message": "Itinerary deleted successfully"}

# Cruise specific endpoints
@api_router.post("/trips/{trip_id}/cruise-info", response_model=CruiseInfo)
async def create_cruise_info(trip_id: str, cruise_data: CruiseInfoCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    cruise_info = CruiseInfo(**cruise_data.dict())
    cruise_dict = prepare_for_mongo(cruise_info.dict())
    
    await db.cruise_info.insert_one(cruise_dict)
    return cruise_info

@api_router.get("/trips/{trip_id}/cruise-info", response_model=Optional[CruiseInfo])
async def get_cruise_info(trip_id: str, current_user: dict = Depends(get_current_user)):
    cruise_info = await db.cruise_info.find_one({"trip_id": trip_id})
    if cruise_info:
        return CruiseInfo(**parse_from_mongo(cruise_info))
    return None

@api_router.put("/cruise-info/{cruise_info_id}", response_model=CruiseInfo)
async def update_cruise_info(cruise_info_id: str, cruise_data: CruiseInfoCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data = prepare_for_mongo(cruise_data.dict())
    await db.cruise_info.update_one({"id": cruise_info_id}, {"$set": update_data})
    
    updated_cruise = await db.cruise_info.find_one({"id": cruise_info_id})
    if not updated_cruise:
        raise HTTPException(status_code=404, detail="Cruise info not found")
    
    return CruiseInfo(**parse_from_mongo(updated_cruise))

@api_router.get("/trips/{trip_id}/port-schedules", response_model=List[PortSchedule])
async def get_port_schedules(trip_id: str, current_user: dict = Depends(get_current_user)):
    schedules = await db.port_schedules.find({"trip_id": trip_id}).to_list(1000)
    return [PortSchedule(**parse_from_mongo(schedule)) for schedule in schedules]

@api_router.post("/port-schedules", response_model=PortSchedule)
async def create_port_schedule(schedule_data: PortScheduleCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    schedule = PortSchedule(**schedule_data.dict())
    schedule_dict = prepare_for_mongo(schedule.dict())
    
    await db.port_schedules.insert_one(schedule_dict)
    return schedule

# Trip Details endpoints
@api_router.get("/trips/{trip_id}/details")
async def get_trip_details(trip_id: str, current_user: dict = Depends(get_current_user)):
    """Get all trip details (cruise, resort, tour, custom) for a specific trip"""
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get the trip to determine type
    trip = await db.trips.find_one({"id": trip_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    result = {"trip_type": trip.get("trip_type", "custom")}
    
    # Get specific details based on trip type
    if trip.get("trip_type") == "cruise":
        cruise_details = await db.cruise_details.find_one({"trip_id": trip_id})
        if cruise_details:
            result["cruise_details"] = CruiseDetails(**parse_from_mongo(cruise_details))
    elif trip.get("trip_type") == "resort":
        resort_details = await db.resort_details.find_one({"trip_id": trip_id})
        if resort_details:
            result["resort_details"] = ResortDetails(**parse_from_mongo(resort_details))
    elif trip.get("trip_type") == "tour":
        tour_details = await db.tour_details.find_one({"trip_id": trip_id})
        if tour_details:
            result["tour_details"] = TourDetails(**parse_from_mongo(tour_details))
    else:
        custom_details = await db.custom_trip_details.find_one({"trip_id": trip_id})
        if custom_details:
            result["custom_details"] = CustomTripDetails(**parse_from_mongo(custom_details))
    
    return result

@api_router.post("/trips/{trip_id}/cruise-details", response_model=CruiseDetails)
async def create_or_update_cruise_details(trip_id: str, details: CruiseDetailsUpdate, current_user: dict = Depends(get_current_user)):
    """Create or update cruise details for a trip"""
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if cruise details already exist
    existing_details = await db.cruise_details.find_one({"trip_id": trip_id})
    
    if existing_details:
        # Update existing details
        update_data = {k: v for k, v in details.dict().items() if v is not None}
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.cruise_details.update_one({"trip_id": trip_id}, {"$set": update_data})
        updated_details = await db.cruise_details.find_one({"trip_id": trip_id})
        return CruiseDetails(**parse_from_mongo(updated_details))
    else:
        # Create new details
        cruise_details = CruiseDetails(trip_id=trip_id, **details.dict(exclude_unset=True))
        cruise_dict = prepare_for_mongo(cruise_details.dict())
        
        await db.cruise_details.insert_one(cruise_dict)
        return cruise_details

@api_router.post("/trips/{trip_id}/resort-details", response_model=ResortDetails)
async def create_or_update_resort_details(trip_id: str, details: ResortDetailsUpdate, current_user: dict = Depends(get_current_user)):
    """Create or update resort details for a trip"""
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if resort details already exist
    existing_details = await db.resort_details.find_one({"trip_id": trip_id})
    
    if existing_details:
        # Update existing details
        update_data = {k: v for k, v in details.dict().items() if v is not None}
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.resort_details.update_one({"trip_id": trip_id}, {"$set": update_data})
        updated_details = await db.resort_details.find_one({"trip_id": trip_id})
        return ResortDetails(**parse_from_mongo(updated_details))
    else:
        # Create new details
        resort_details = ResortDetails(trip_id=trip_id, **details.dict(exclude_unset=True))
        resort_dict = prepare_for_mongo(resort_details.dict())
        
        await db.resort_details.insert_one(resort_dict)
        return resort_details

@api_router.post("/trips/{trip_id}/tour-details", response_model=TourDetails)
async def create_or_update_tour_details(trip_id: str, details: TourDetailsUpdate, current_user: dict = Depends(get_current_user)):
    """Create or update tour details for a trip"""
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if tour details already exist
    existing_details = await db.tour_details.find_one({"trip_id": trip_id})
    
    if existing_details:
        # Update existing details
        update_data = {k: v for k, v in details.dict().items() if v is not None}
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.tour_details.update_one({"trip_id": trip_id}, {"$set": update_data})
        updated_details = await db.tour_details.find_one({"trip_id": trip_id})
        return TourDetails(**parse_from_mongo(updated_details))
    else:
        # Create new details
        tour_details = TourDetails(trip_id=trip_id, **details.dict(exclude_unset=True))
        tour_dict = prepare_for_mongo(tour_details.dict())
        
        await db.tour_details.insert_one(tour_dict)
        return tour_details

@api_router.post("/trips/{trip_id}/custom-details", response_model=CustomTripDetails)
async def create_or_update_custom_details(trip_id: str, details: CustomTripDetailsUpdate, current_user: dict = Depends(get_current_user)):
    """Create or update custom trip details for a trip"""
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if custom details already exist
    existing_details = await db.custom_trip_details.find_one({"trip_id": trip_id})
    
    if existing_details:
        # Update existing details
        update_data = {k: v for k, v in details.dict().items() if v is not None}
        update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
        
        await db.custom_trip_details.update_one({"trip_id": trip_id}, {"$set": update_data})
        updated_details = await db.custom_trip_details.find_one({"trip_id": trip_id})
        return CustomTripDetails(**parse_from_mongo(updated_details))
    else:
        # Create new details
        custom_details = CustomTripDetails(trip_id=trip_id, **details.dict(exclude_unset=True))
        custom_dict = prepare_for_mongo(custom_details.dict())
        
        await db.custom_trip_details.insert_one(custom_dict)
        return custom_details

# POI endpoints
@api_router.get("/pois", response_model=List[POI])
async def get_pois(category: Optional[POICategory] = None, current_user: dict = Depends(get_current_user)):
    query = {}
    if category:
        query["category"] = category
    
    pois = await db.pois.find(query).to_list(1000)
    return [POI(**parse_from_mongo(poi)) for poi in pois]

@api_router.post("/pois", response_model=POI)
async def create_poi(poi_data: POICreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    poi = POI(**poi_data.dict())
    poi_dict = prepare_for_mongo(poi.dict())
    
    await db.pois.insert_one(poi_dict)
    return poi

# Photo endpoints
@api_router.post("/trips/{trip_id}/photos")
async def upload_photo(
    trip_id: str,
    file: UploadFile = File(...),
    caption: str = Form(""),
    photo_category: PhotoCategory = Form(...),
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "client":
        # Check if agent owns this trip
        trip = await db.trips.find_one({"id": trip_id})
        if not trip or (current_user["role"] == "agent" and trip["agent_id"] != current_user["id"]):
            raise HTTPException(status_code=403, detail="Not authorized")
    else:
        # Check if client owns this trip
        trip = await db.trips.find_one({"id": trip_id})
        if not trip or trip["client_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Not authorized")
    
    # Create uploads directory if it doesn't exist
    upload_dir = Path("uploads")
    upload_dir.mkdir(exist_ok=True)
    
    # Save file
    file_id = str(uuid.uuid4())
    file_extension = file.filename.split(".")[-1] if "." in file.filename else "jpg"
    filename = f"{file_id}.{file_extension}"
    file_path = upload_dir / filename
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    
    # Save photo info to database
    photo = ClientPhoto(
        trip_id=trip_id,
        client_id=current_user["id"],
        url=f"/uploads/{filename}",
        caption=caption,
        photo_category=photo_category
    )
    
    photo_dict = prepare_for_mongo(photo.dict())
    await db.client_photos.insert_one(photo_dict)
    
    return photo

@api_router.get("/trips/{trip_id}/photos", response_model=List[ClientPhoto])
async def get_trip_photos(
    trip_id: str,
    category: Optional[PhotoCategory] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {"trip_id": trip_id}
    if category:
        query["photo_category"] = category
    
    photos = await db.client_photos.find(query).to_list(1000)
    return [ClientPhoto(**parse_from_mongo(photo)) for photo in photos]

# Client notes endpoints - Fixed for Admin/Agent visibility
@api_router.get("/trips/{trip_id}/notes", response_model=List[ClientNote])
async def get_client_notes(trip_id: str, current_user: dict = Depends(get_current_user)):
    # Admin and agents can see all notes for a trip, clients see only their own
    query = {"trip_id": trip_id}
    if current_user["role"] == "client":
        query["client_id"] = current_user["id"]
    
    notes = await db.client_notes.find(query).to_list(1000)
    return [ClientNote(**parse_from_mongo(note)) for note in notes]

@api_router.post("/trips/{trip_id}/notes", response_model=ClientNote)
async def create_client_note(trip_id: str, note_data: ClientNoteCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "client":
        raise HTTPException(status_code=403, detail="Only clients can create notes")
    
    note = ClientNote(**note_data.dict(), client_id=current_user["id"])
    note_dict = prepare_for_mongo(note.dict())
    
    await db.client_notes.insert_one(note_dict)
    return note

@api_router.put("/notes/{note_id}")
async def update_client_note(
    note_id: str, 
    note_data: dict, 
    current_user: dict = Depends(get_current_user)
):
    # Find the note
    note = await db.client_notes.find_one({"id": note_id})
    if not note:
        raise HTTPException(status_code=404, detail="Note not found")
    
    # Authorization: client can edit their own notes, admin/agent can edit any
    if current_user["role"] == "client":
        if note["client_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Not authorized to edit this note")
    elif current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_data = {
        "note_text": note_data.get("note_text", note["note_text"]),
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    
    await db.client_notes.update_one({"id": note_id}, {"$set": update_data})
    updated_note = await db.client_notes.find_one({"id": note_id})
    
    return {"message": "Note updated successfully", "note": ClientNote(**parse_from_mongo(updated_note))}

# Get all notes for admin/agent dashboard
@api_router.get("/notes/all")
async def get_all_client_notes(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get all notes with trip and client info
    notes = await db.client_notes.find().to_list(1000)
    enriched_notes = []
    
    for note in notes:
        # Get trip info
        trip = await db.trips.find_one({"id": note["trip_id"]})
        # Get client info
        client = await db.users.find_one({"id": note["client_id"]})
        
        enriched_note = parse_from_mongo(note)
        enriched_note["trip_title"] = trip["title"] if trip else "Unknown Trip"
        enriched_note["client_name"] = f"{client['first_name']} {client['last_name']}" if client else "Unknown Client"
        
        enriched_notes.append(enriched_note)
    
    return enriched_notes

# Users management (admin only)
@api_router.get("/users", response_model=List[User])
async def get_users(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    users_data = await db.users.find().to_list(1000)
    
    # If agent, only show clients they can manage
    if current_user["role"] == "agent":
        users_data = [user for user in users_data if user.get("role") == "client"]
    
    return [User(**parse_from_mongo(user)) for user in users_data]

@api_router.get("/users/{user_id}", response_model=User)
async def get_user_by_id(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    user_data = await db.users.find_one({"id": user_id})
    if not user_data:
        raise HTTPException(status_code=404, detail="User not found")
    
    # If agent, verify they can access this user (only clients)
    if current_user["role"] == "agent":
        if user_data.get("role") != "client":
            raise HTTPException(status_code=403, detail="Not authorized to access this user")
    
    return User(**parse_from_mongo(user_data))

# Clients management (admin and agent)
@api_router.get("/clients", response_model=List[User])
async def get_clients(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get all clients
    clients = await db.users.find({"role": "client"}).to_list(1000)
    return [User(**parse_from_mongo(client)) for client in clients]

@api_router.put("/users/{user_id}", response_model=User)
async def update_user(user_id: str, user_data: UserUpdate, current_user: dict = Depends(get_current_user)):
    # Check permissions
    user_to_update = await db.users.find_one({"id": user_id})
    if not user_to_update:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Permission checks
    if current_user["role"] == "agent":
        # Agents can only update clients
        if user_to_update["role"] != "client":
            raise HTTPException(status_code=403, detail="Agents can only update clients")
    elif current_user["role"] == "admin":
        # Admins can update anyone
        pass
    else:
        # Clients can only update themselves (not implemented here)
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Update user
    update_data = {k: v for k, v in user_data.dict(exclude_unset=True).items() if v is not None}
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
    
    updated_user = await db.users.find_one({"id": user_id})
    return User(**parse_from_mongo(updated_user))

@api_router.post("/users/{user_id}/block")
async def block_user(user_id: str, current_user: dict = Depends(get_current_user)):
    user_to_block = await db.users.find_one({"id": user_id})
    if not user_to_block:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Permission checks
    if current_user["role"] == "agent":
        # Agents can only block clients
        if user_to_block["role"] != "client":
            raise HTTPException(status_code=403, detail="Agents can only block clients")
    elif current_user["role"] == "admin":
        # Admins can block anyone except other admins
        if user_to_block["role"] == "admin" and user_to_block["id"] != current_user["id"]:
            # Don't allow blocking other admins (prevent lockout)
            if user_to_block["id"] != current_user["id"]:
                pass  # Allow for now, but could add more restrictions
    else:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.users.update_one({"id": user_id}, {"$set": {"blocked": True}})
    return {"message": "User blocked successfully"}

@api_router.post("/users/{user_id}/unblock")
async def unblock_user(user_id: str, current_user: dict = Depends(get_current_user)):
    user_to_unblock = await db.users.find_one({"id": user_id})
    if not user_to_unblock:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Permission checks (same as block)
    if current_user["role"] == "agent":
        if user_to_unblock["role"] != "client":
            raise HTTPException(status_code=403, detail="Agents can only unblock clients")
    elif current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Not authorized")
    
    await db.users.update_one({"id": user_id}, {"$set": {"blocked": False}})
    return {"message": "User unblocked successfully"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    
    user_to_delete = await db.users.find_one({"id": user_id})
    if not user_to_delete:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Don't allow deleting yourself
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    
    return {"message": "User deleted successfully"}

# Dashboard stats
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "admin":
        total_trips = await db.trips.count_documents({})
        total_users = await db.users.count_documents({})
        confirmed_trips = await db.trips.count_documents({"status": "confirmed"})
        
        return {
            "total_trips": total_trips,
            "total_users": total_users,
            "confirmed_trips": confirmed_trips
        }
    elif current_user["role"] == "agent":
        agent_trips = await db.trips.count_documents({"agent_id": current_user["id"]})
        confirmed_trips = await db.trips.count_documents({"agent_id": current_user["id"], "status": "confirmed"})
        
        return {
            "my_trips": agent_trips,
            "confirmed_trips": confirmed_trips,
            "completed_trips": await db.trips.count_documents({"agent_id": current_user["id"], "status": "completed"})
        }
    else:  # client
        my_trips = await db.trips.count_documents({"client_id": current_user["id"]})
        my_photos = await db.client_photos.count_documents({"client_id": current_user["id"]})
        
        return {
            "my_trips": my_trips,
            "my_photos": my_photos,
            "upcoming_trips": await db.trips.count_documents({
                "client_id": current_user["id"],
                "start_date": {"$gte": datetime.now(timezone.utc).isoformat()}
            })
        }

# Trip Administration endpoints (Admin/Agent only)
@api_router.post("/trips/{trip_id}/admin", response_model=TripAdmin)
async def create_trip_admin(trip_id: str, admin_data: TripAdminCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if trip exists and user has access
    trip = await db.trips.find_one({"id": trip_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    if current_user["role"] == "agent" and trip["agent_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized to manage this trip")
    
    # Calculate derived fields
    admin_dict = prepare_for_mongo(admin_data.dict())
    calculated_data = calculate_trip_admin_fields(admin_dict)
    
    trip_admin = TripAdmin(**calculated_data)
    admin_dict = prepare_for_mongo(trip_admin.dict())
    
    await db.trip_admin.insert_one(admin_dict)
    return trip_admin

@api_router.get("/trips/{trip_id}/admin", response_model=Optional[TripAdmin])
async def get_trip_admin(trip_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    trip_admin = await db.trip_admin.find_one({"trip_id": trip_id})
    if trip_admin:
        # Get installments and recalculate
        installments = await db.payment_installments.find({"trip_admin_id": trip_admin["id"]}).to_list(1000)
        calculated_data = calculate_trip_admin_fields(trip_admin, installments)
        return TripAdmin(**parse_from_mongo(calculated_data))
    
    return None

@api_router.put("/trip-admin/{admin_id}", response_model=TripAdmin)
async def update_trip_admin(admin_id: str, admin_data: TripAdminUpdate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    existing = await db.trip_admin.find_one({"id": admin_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Trip admin not found")
    
    # Update fields
    update_data = {k: v for k, v in admin_data.dict(exclude_unset=True).items() if v is not None}
    update_data["updated_at"] = datetime.now(timezone.utc).isoformat()
    
    merged_data = {**existing, **prepare_for_mongo(update_data)}
    
    # Get installments and recalculate
    installments = await db.payment_installments.find({"trip_admin_id": admin_id}).to_list(1000)
    calculated_data = calculate_trip_admin_fields(merged_data, installments)
    
    await db.trip_admin.update_one({"id": admin_id}, {"$set": calculated_data})
    updated_admin = await db.trip_admin.find_one({"id": admin_id})
    
    return TripAdmin(**parse_from_mongo(updated_admin))

# Payment Installments endpoints
@api_router.post("/trip-admin/{admin_id}/payments", response_model=PaymentInstallment)
async def create_payment_installment(admin_id: str, payment_data: PaymentInstallmentCreate, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    payment = PaymentInstallment(**payment_data.dict())
    payment_dict = prepare_for_mongo(payment.dict())
    
    await db.payment_installments.insert_one(payment_dict)
    
    # Recalculate trip admin balance
    installments = await db.payment_installments.find({"trip_admin_id": admin_id}).to_list(1000)
    trip_admin = await db.trip_admin.find_one({"id": admin_id})
    
    if trip_admin:
        calculated_data = calculate_trip_admin_fields(trip_admin, installments)
        await db.trip_admin.update_one({"id": admin_id}, {"$set": {"balance_due": calculated_data["balance_due"]}})
    
    return payment

@api_router.get("/trip-admin/{admin_id}/payments", response_model=List[PaymentInstallment])
async def get_payment_installments(admin_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    payments = await db.payment_installments.find({"trip_admin_id": admin_id}).to_list(1000)
    return [PaymentInstallment(**parse_from_mongo(payment)) for payment in payments]

@api_router.delete("/payments/{payment_id}")
async def delete_payment_installment(payment_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get payment to find admin_id for recalculation
    payment = await db.payment_installments.find_one({"id": payment_id})
    if not payment:
        raise HTTPException(status_code=404, detail="Payment not found")
    
    admin_id = payment["trip_admin_id"]
    
    # Delete payment
    await db.payment_installments.delete_one({"id": payment_id})
    
    # Recalculate trip admin balance
    installments = await db.payment_installments.find({"trip_admin_id": admin_id}).to_list(1000)
    trip_admin = await db.trip_admin.find_one({"id": admin_id})
    
    if trip_admin:
        calculated_data = calculate_trip_admin_fields(trip_admin, installments)
        await db.trip_admin.update_one({"id": admin_id}, {"$set": {"balance_due": calculated_data["balance_due"]}})
    
    return {"message": "Payment deleted successfully"}

# Financial Analytics endpoints
@api_router.get("/analytics/agent-commissions")
async def get_agent_commission_analytics(
    year: int = None,
    agent_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # If agent, can only see own data
    if current_user["role"] == "agent":
        agent_id = current_user["id"]
    
    # Build query
    query = {}
    if agent_id:
        # Get trips for this agent
        agent_trips = await db.trips.find({"agent_id": agent_id}).to_list(1000)
        trip_ids = [trip["id"] for trip in agent_trips]
        query["trip_id"] = {"$in": trip_ids}
    
    if year:
        start_date = datetime(year, 1, 1, tzinfo=timezone.utc)
        end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        query["practice_confirm_date"] = {
            "$gte": start_date.isoformat(),
            "$lt": end_date.isoformat()
        }
    
    # Get confirmed trip admin records
    query["status"] = "confirmed"
    confirmed_trips = await db.trip_admin.find(query).to_list(1000)
    
    # Parse MongoDB data to remove ObjectIds
    parsed_trips = [parse_from_mongo(trip) for trip in confirmed_trips]
    
    # Calculate totals
    total_revenue = sum(trip.get("gross_amount", 0) for trip in confirmed_trips)
    total_gross_commission = sum(trip.get("gross_commission", 0) for trip in confirmed_trips)
    total_supplier_commission = sum(trip.get("supplier_commission", 0) for trip in confirmed_trips)
    total_agent_commission = sum(trip.get("agent_commission", 0) for trip in confirmed_trips)
    total_discounts = sum(trip.get("discount", 0) for trip in confirmed_trips)
    
    return {
        "year": year or "all_time",
        "agent_id": agent_id,
        "total_confirmed_trips": len(confirmed_trips),
        "total_revenue": total_revenue,
        "total_gross_commission": total_gross_commission,
        "total_supplier_commission": total_supplier_commission,
        "total_agent_commission": total_agent_commission,
        "total_discounts": total_discounts,
        "trips": parsed_trips
    }

# Complete Financial Reports with monthly/annual breakdowns
@api_router.get("/reports/financial")
async def get_financial_reports(
    year: int = None,
    month: int = None,
    agent_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # If agent, can only see own data
    if current_user["role"] == "agent":
        agent_id = current_user["id"]
    
    # Build date query
    date_query = {}
    if year:
        start_date = datetime(year, 1, 1, tzinfo=timezone.utc)
        if month:
            # Specific month
            if month == 12:
                end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
            else:
                end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
            start_date = datetime(year, month, 1, tzinfo=timezone.utc)
        else:
            # Entire year
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        
        date_query = {
            "$gte": start_date.isoformat(),
            "$lt": end_date.isoformat()
        }
    
    # Build main query
    query = {"status": "confirmed"}
    if date_query:
        query["practice_confirm_date"] = date_query
    
    if agent_id:
        agent_trips = await db.trips.find({"agent_id": agent_id}).to_list(1000)
        trip_ids = [trip["id"] for trip in agent_trips]
        query["trip_id"] = {"$in": trip_ids}
    
    # Get confirmed trips
    confirmed_trips = await db.trip_admin.find(query).to_list(1000)
    parsed_trips = [parse_from_mongo(trip) for trip in confirmed_trips]
    
    # Enrich trips with client and agent information
    enriched_trips = []
    for trip_admin in parsed_trips:
        # Get trip details
        trip = await db.trips.find_one({"id": trip_admin.get("trip_id")})
        if trip:
            # Get client info
            client = await db.users.find_one({"id": trip.get("client_id")})
            # Get agent info
            agent = await db.users.find_one({"id": trip.get("agent_id")})
            
            # Enrich trip_admin data with names
            enriched_trip = trip_admin.copy()
            enriched_trip["trip_title"] = trip.get("title", "Unknown Trip")
            enriched_trip["trip_destination"] = trip.get("destination", "Unknown Destination")
            enriched_trip["client_name"] = f"{client.get('first_name', 'Unknown')} {client.get('last_name', 'Client')}" if client else "Unknown Client"
            enriched_trip["agent_name"] = f"{agent.get('first_name', 'Unknown')} {agent.get('last_name', 'Agent')}" if agent else "Unknown Agent"
            enriched_trip["client_email"] = client.get("email", "") if client else ""
            enriched_trip["agent_email"] = agent.get("email", "") if agent else ""
            
            enriched_trips.append(enriched_trip)
        else:
            # Keep original if trip not found (shouldn't happen with proper data)
            enriched_trips.append(trip_admin)
    
    # Calculate monthly breakdowns if year provided but no specific month
    monthly_data = []
    if year and not month:
        for m in range(1, 13):
            month_start = datetime(year, m, 1, tzinfo=timezone.utc)
            if m == 12:
                month_end = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
            else:
                month_end = datetime(year, m + 1, 1, tzinfo=timezone.utc)
            
            month_query = query.copy()
            month_query["practice_confirm_date"] = {
                "$gte": month_start.isoformat(),
                "$lt": month_end.isoformat()
            }
            
            month_trips = await db.trip_admin.find(month_query).to_list(1000)
            
            monthly_summary = {
                "month": m,
                "month_name": month_start.strftime("%B"),
                "total_trips": len(month_trips),
                "gross_revenue": sum(trip.get("gross_amount", 0) for trip in month_trips),
                "total_discounts": sum(trip.get("discount", 0) for trip in month_trips),
                "supplier_commissions": sum(trip.get("supplier_commission", 0) for trip in month_trips),
                "agent_commissions": sum(trip.get("agent_commission", 0) for trip in month_trips),
                "client_departures": len(set(trip.get("trip_id", "") for trip in month_trips))
            }
            monthly_data.append(monthly_summary)
    
    # Calculate totals using enriched_trips
    totals = {
        "total_trips": len(enriched_trips),
        "gross_revenue": sum(trip.get("gross_amount", 0) for trip in enriched_trips),
        "total_discounts": sum(trip.get("discount", 0) for trip in enriched_trips),
        "supplier_commissions": sum(trip.get("supplier_commission", 0) for trip in enriched_trips),
        "agent_commissions": sum(trip.get("agent_commission", 0) for trip in enriched_trips),
        "net_revenue": sum(trip.get("net_amount", 0) for trip in enriched_trips),
        "client_departures": len(set(trip.get("trip_id", "") for trip in enriched_trips))
    }
    
    return {
        "period": {
            "year": year,
            "month": month,
            "agent_id": agent_id
        },
        "totals": totals,
        "monthly_breakdown": monthly_data,
        "detailed_trips": enriched_trips,
        "can_export_excel": current_user["role"] == "admin"  # Only admin can export to Excel
    }

# Excel Export endpoint
@api_router.get("/reports/financial/export")
async def export_financial_report(
    year: int = None,
    month: int = None,
    agent_id: str = None,
    current_user: dict = Depends(get_current_user)
):
    from fastapi.responses import StreamingResponse
    import io
    
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admins can export Excel reports")
    
    # Reuse the same logic from get_financial_reports
    # Build date query
    date_query = {}
    if year:
        start_date = datetime(year, 1, 1, tzinfo=timezone.utc)
        if month:
            # Specific month
            if month == 12:
                end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
            else:
                end_date = datetime(year, month + 1, 1, tzinfo=timezone.utc)
            start_date = datetime(year, month, 1, tzinfo=timezone.utc)
        else:
            # Entire year
            end_date = datetime(year + 1, 1, 1, tzinfo=timezone.utc)
        
        date_query = {
            "$gte": start_date.isoformat(),
            "$lt": end_date.isoformat()
        }
    
    # Build main query
    query = {"status": "confirmed"}
    if date_query:
        query["practice_confirm_date"] = date_query
    
    if agent_id:
        agent_trips = await db.trips.find({"agent_id": agent_id}).to_list(1000)
        trip_ids = [trip["id"] for trip in agent_trips]
        query["trip_id"] = {"$in": trip_ids}
    
    # Get confirmed trips
    confirmed_trips = await db.trip_admin.find(query).to_list(1000)
    
    # Enrich trips with client and agent information (same logic as reports)
    enriched_trips = []
    for trip_admin in confirmed_trips:
        trip = await db.trips.find_one({"id": trip_admin.get("trip_id")})
        if trip:
            client = await db.users.find_one({"id": trip.get("client_id")})
            agent = await db.users.find_one({"id": trip.get("agent_id")})
            
            enriched_trip = {
                "practice_number": trip_admin.get("practice_number", ""),
                "booking_number": trip_admin.get("booking_number", ""),
                "client_name": f"{client.get('first_name', 'Unknown')} {client.get('last_name', 'Client')}" if client else "Unknown Client",
                "practice_confirm_date": trip_admin.get("practice_confirm_date", ""),
                "client_departure_date": trip_admin.get("client_departure_date", ""),
                "gross_amount": trip_admin.get("gross_amount", 0),
                "supplier_commission": trip_admin.get("supplier_commission", 0),
                "discount": trip_admin.get("discount", 0),
                "agent_commission": trip_admin.get("agent_commission", 0)
            }
            enriched_trips.append(enriched_trip)
    
    # Create Excel file using openpyxl
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=500, detail="Excel export library not available")
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Report Finanziario"
    
    # Header row
    headers = [
        "Numero Pratica",
        "Numero Prenotazione", 
        "Nome Cliente",
        "Data Conferma Pratica",
        "Data Partenza",
        "Fatturato Lordo (€)",
        "Commissione Fornitore (€)",
        "Sconti Applicati (€)",
        "Commissione Agente (€)"
    ]
    
    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row_idx, trip in enumerate(enriched_trips, 2):
        # Format dates
        confirm_date = ""
        departure_date = ""
        
        if trip["practice_confirm_date"]:
            try:
                confirm_date = datetime.fromisoformat(trip["practice_confirm_date"].replace('Z', '+00:00')).strftime('%d/%m/%Y')
            except:
                confirm_date = trip["practice_confirm_date"]
        
        if trip["client_departure_date"]:
            try:
                departure_date = datetime.fromisoformat(trip["client_departure_date"].replace('Z', '+00:00')).strftime('%d/%m/%Y')
            except:
                departure_date = trip["client_departure_date"]
        
        row_data = [
            trip["practice_number"],
            trip["booking_number"],
            trip["client_name"],
            confirm_date,
            departure_date,
            trip["gross_amount"],
            trip["supplier_commission"],
            trip["discount"],
            trip["agent_commission"]
        ]
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        max_length = 0
        column = get_column_letter(col)
        
        for row in ws[column]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to BytesIO
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    # Create filename
    filename = "report_finanziario"
    if year and month:
        filename += f"_{year}_{month:02d}"
    elif year:
        filename += f"_{year}"
    else:
        filename += "_tutti_anni"
    filename += ".xlsx"
    
    # Return file
    return StreamingResponse(
        io.BytesIO(excel_buffer.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

# Financial Sheets Management
@api_router.post("/financial-sheets")
async def create_financial_sheet(
    sheet_data: dict,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    sheet = {
        "id": str(uuid.uuid4()),
        "created_by": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "title": sheet_data.get("title", "Scheda Finanziaria"),
        "description": sheet_data.get("description", ""),
        "year": sheet_data.get("year"),
        "month": sheet_data.get("month"),
        "agent_id": sheet_data.get("agent_id"),
        "data": sheet_data.get("data", {}),
        "status": "draft"
    }
    
    await db.financial_sheets.insert_one(sheet)
    return {"message": "Financial sheet created successfully", "sheet_id": sheet["id"]}

@api_router.get("/financial-sheets")
async def get_financial_sheets(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    query = {}
    if current_user["role"] == "agent":
        query = {"$or": [{"created_by": current_user["id"]}, {"agent_id": current_user["id"]}]}
    
    sheets = await db.financial_sheets.find(query).to_list(1000)
    return [parse_from_mongo(sheet) for sheet in sheets]

@api_router.put("/financial-sheets/{sheet_id}")
async def update_financial_sheet(
    sheet_id: str,
    sheet_data: dict,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Check if sheet exists and user has permission
    sheet = await db.financial_sheets.find_one({"id": sheet_id})
    if not sheet:
        raise HTTPException(status_code=404, detail="Financial sheet not found")
    
    if current_user["role"] == "agent":
        if sheet["created_by"] != current_user["id"] and sheet.get("agent_id") != current_user["id"]:
            raise HTTPException(status_code=403, detail="Not authorized to update this sheet")
    
    update_data = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    update_data.update(sheet_data)
    
    await db.financial_sheets.update_one({"id": sheet_id}, {"$set": update_data})
    return {"message": "Financial sheet updated successfully"}

# Trip Status Management - Fix for draft status issue
@api_router.put("/trips/{trip_id}/status")
async def update_trip_status(
    trip_id: str,
    status_data: dict,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    trip = await db.trips.find_one({"id": trip_id})
    if not trip:
        raise HTTPException(status_code=404, detail="Trip not found")
    
    # Check if agent owns this trip
    if current_user["role"] == "agent" and trip["agent_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized to update this trip")
    
    new_status = status_data.get("status")
    if new_status not in ["draft", "active", "confirmed", "completed", "cancelled"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    update_data = {
        "status": new_status,
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    
    # If activating or confirming a trip, ensure it has minimum required data
    if new_status in ["active", "confirmed"]:
        if not trip.get("title") or not trip.get("client_id"):
            raise HTTPException(status_code=400, detail="Trip must have title and client to be activated/confirmed")
    
    # Update trip status
    await db.trips.update_one({"id": trip_id}, {"$set": update_data})
    
    # IMPORTANT: When trip is confirmed, automatically confirm its administrative data
    if new_status == "confirmed":
        # Find and update trip_admin status to "confirmed"
        trip_admin = await db.trip_admin.find_one({"trip_id": trip_id})
        if trip_admin:
            admin_update_data = {
                "status": "confirmed",
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user["id"]
            }
            await db.trip_admin.update_one({"trip_id": trip_id}, {"$set": admin_update_data})
            print(f"✅ Trip {trip_id} confirmed - Administrative data also confirmed")
        else:
            print(f"⚠️ Trip {trip_id} confirmed but no administrative data found")
    
    # If trip is moved back to draft, also set admin data to draft
    elif new_status == "draft":
        trip_admin = await db.trip_admin.find_one({"trip_id": trip_id})
        if trip_admin:
            admin_update_data = {
                "status": "draft",
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "updated_by": current_user["id"]
            }
            await db.trip_admin.update_one({"trip_id": trip_id}, {"$set": admin_update_data})
            print(f"📝 Trip {trip_id} moved to draft - Administrative data also set to draft")
    
    return {"message": f"Trip status updated to {new_status}"}

# Quote Request Feature
@api_router.post("/quote-requests")
async def create_quote_request(
    request_data: dict,
    current_user: dict = Depends(get_current_user)
):
    if current_user["role"] != "client":
        raise HTTPException(status_code=403, detail="Only clients can create quote requests")
    
    quote_request = {
        "id": str(uuid.uuid4()),
        "client_id": current_user["id"],
        "created_at": datetime.now(timezone.utc).isoformat(),
        "destination": request_data.get("destination", ""),
        "travel_dates": request_data.get("travel_dates", ""),
        "number_of_travelers": request_data.get("number_of_travelers", 1),
        "trip_type": request_data.get("trip_type", "custom"),
        "budget_range": request_data.get("budget_range", ""),
        "special_requirements": request_data.get("special_requirements", ""),
        "contact_preference": request_data.get("contact_preference", "email"),
        "status": "pending",
        "notes": request_data.get("notes", "")
    }
    
    await db.quote_requests.insert_one(quote_request)
    
    return {
        "message": "Quote request submitted successfully",
        "request_id": quote_request["id"]
    }

@api_router.get("/quote-requests")
async def get_quote_requests(current_user: dict = Depends(get_current_user)):
    if current_user["role"] == "client":
        # Clients see only their own requests
        requests = await db.quote_requests.find({"client_id": current_user["id"]}).to_list(1000)
    elif current_user["role"] in ["admin", "agent"]:
        # Admin and agents see all requests
        requests = await db.quote_requests.find().to_list(1000)
    else:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    return [parse_from_mongo(request) for request in requests]

@api_router.put("/quote-requests/{request_id}")
async def update_quote_request(
    request_id: str,
    update_data: dict,
    current_user: dict = Depends(get_current_user)
):
    request = await db.quote_requests.find_one({"id": request_id})
    if not request:
        raise HTTPException(status_code=404, detail="Quote request not found")
    
    # Authorization check
    if current_user["role"] == "client":
        if request["client_id"] != current_user["id"]:
            raise HTTPException(status_code=403, detail="Not authorized")
    elif current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    update_dict = {
        "updated_at": datetime.now(timezone.utc).isoformat(),
        "updated_by": current_user["id"]
    }
    update_dict.update(update_data)
    
    await db.quote_requests.update_one({"id": request_id}, {"$set": update_dict})
    return {"message": "Quote request updated successfully"}

# Client Details and Financial Summary
@api_router.get("/clients/{client_id}/details")
async def get_client_details(client_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get client info
    client = await db.users.find_one({"id": client_id, "role": "client"})
    if not client:
        raise HTTPException(status_code=404, detail="Client not found")
    
    # Get client's trips
    client_trips = await db.trips.find({"client_id": client_id}).to_list(1000)
    
    # Get financial data for each trip
    trips_with_financials = []
    for trip in client_trips:
        # Get trip admin data (financial details)
        trip_admin = await db.trip_admin.find_one({"trip_id": trip["id"]})
        
        trip_data = parse_from_mongo(trip)
        if trip_admin:
            trip_data["financial"] = parse_from_mongo(trip_admin)
        else:
            trip_data["financial"] = None
            
        trips_with_financials.append(trip_data)
    
    return {
        "client": parse_from_mongo(client),
        "trips": trips_with_financials
    }

@api_router.get("/clients/{client_id}/financial-summary")
async def get_client_financial_summary(client_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get all confirmed trips for this client
    client_trips = await db.trips.find({"client_id": client_id}).to_list(1000)
    trip_ids = [trip["id"] for trip in client_trips]
    
    # Get confirmed financial data
    confirmed_financials = await db.trip_admin.find({
        "trip_id": {"$in": trip_ids}, 
        "status": "confirmed"
    }).to_list(1000)
    
    # Calculate totals
    total_gross_amount = sum(f.get("gross_amount", 0) for f in confirmed_financials)
    total_net_amount = sum(f.get("net_amount", 0) for f in confirmed_financials)
    total_discounts = sum(f.get("discount", 0) for f in confirmed_financials)
    total_supplier_commission = sum(f.get("supplier_commission", 0) for f in confirmed_financials)
    total_agent_commission = sum(f.get("agent_commission", 0) for f in confirmed_financials)
    total_confirmations = len(confirmed_financials)
    
    # Get pending trips (those with admin data but not confirmed)
    pending_financials = await db.trip_admin.find({
        "trip_id": {"$in": trip_ids}, 
        "status": {"$ne": "confirmed"}
    }).to_list(1000)
    
    pending_gross_amount = sum(f.get("gross_amount", 0) for f in pending_financials)
    pending_count = len(pending_financials)
    
    # Get trips without financial data
    trips_with_admin = set()
    for f in confirmed_financials + pending_financials:
        trips_with_admin.add(f.get("trip_id"))
    
    trips_without_financial = len([t for t in client_trips if t["id"] not in trips_with_admin])
    
    return {
        "client_id": client_id,
        "confirmed_bookings": {
            "count": total_confirmations,
            "total_gross_amount": total_gross_amount,
            "total_net_amount": total_net_amount,
            "total_discounts": total_discounts,
            "total_supplier_commission": total_supplier_commission,
            "total_agent_commission": total_agent_commission
        },
        "pending_bookings": {
            "count": pending_count,
            "pending_gross_amount": pending_gross_amount
        },
        "stats": {
            "total_trips": len(client_trips),
            "trips_without_financial_data": trips_without_financial,
            "average_trip_value": total_gross_amount / total_confirmations if total_confirmations > 0 else 0
        }
    }

# Cleanup endpoint for orphaned data
@api_router.post("/admin/cleanup-orphaned-data")
async def cleanup_orphaned_data(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Only admins can cleanup orphaned data")
    
    print("🧹 Starting cleanup of orphaned data...")
    
    # Get all existing trip IDs
    existing_trips = await db.trips.find({}, {"id": 1}).to_list(10000)
    existing_trip_ids = set(trip["id"] for trip in existing_trips)
    print(f"   📊 Found {len(existing_trip_ids)} existing trips")
    
    cleanup_results = {}
    
    # Clean orphaned trip_admin records
    all_admin_records = await db.trip_admin.find({}).to_list(10000)
    orphaned_admin = [record for record in all_admin_records if record.get("trip_id") not in existing_trip_ids]
    if orphaned_admin:
        orphaned_admin_ids = [record["trip_id"] for record in orphaned_admin]
        admin_result = await db.trip_admin.delete_many({"trip_id": {"$in": orphaned_admin_ids}})
        cleanup_results["trip_admin"] = admin_result.deleted_count
        print(f"   🗑️ Deleted {admin_result.deleted_count} orphaned financial records")
    else:
        cleanup_results["trip_admin"] = 0
    
    # Clean orphaned itineraries
    all_itineraries = await db.itineraries.find({}).to_list(10000)
    orphaned_itineraries = [record for record in all_itineraries if record.get("trip_id") not in existing_trip_ids]
    if orphaned_itineraries:
        orphaned_itinerary_ids = [record["trip_id"] for record in orphaned_itineraries]
        itinerary_result = await db.itineraries.delete_many({"trip_id": {"$in": orphaned_itinerary_ids}})
        cleanup_results["itineraries"] = itinerary_result.deleted_count
        print(f"   🗑️ Deleted {itinerary_result.deleted_count} orphaned itinerary records")
    else:
        cleanup_results["itineraries"] = 0
    
    # Clean orphaned cruise_info
    all_cruise_info = await db.cruise_info.find({}).to_list(10000)
    orphaned_cruise = [record for record in all_cruise_info if record.get("trip_id") not in existing_trip_ids]
    if orphaned_cruise:
        orphaned_cruise_ids = [record["trip_id"] for record in orphaned_cruise]
        cruise_result = await db.cruise_info.delete_many({"trip_id": {"$in": orphaned_cruise_ids}})
        cleanup_results["cruise_info"] = cruise_result.deleted_count
        print(f"   🗑️ Deleted {cruise_result.deleted_count} orphaned cruise info records")
    else:
        cleanup_results["cruise_info"] = 0
    
    # Clean orphaned client_notes
    all_notes = await db.client_notes.find({}).to_list(10000)
    orphaned_notes = [record for record in all_notes if record.get("trip_id") not in existing_trip_ids]
    if orphaned_notes:
        orphaned_note_ids = [record["trip_id"] for record in orphaned_notes]
        notes_result = await db.client_notes.delete_many({"trip_id": {"$in": orphaned_note_ids}})
        cleanup_results["client_notes"] = notes_result.deleted_count
        print(f"   🗑️ Deleted {notes_result.deleted_count} orphaned client notes")
    else:
        cleanup_results["client_notes"] = 0
    
    # Clean orphaned client_photos
    all_photos = await db.client_photos.find({}).to_list(10000)
    orphaned_photos = [record for record in all_photos if record.get("trip_id") not in existing_trip_ids]
    if orphaned_photos:
        orphaned_photo_ids = [record["trip_id"] for record in orphaned_photos]
        photos_result = await db.client_photos.delete_many({"trip_id": {"$in": orphaned_photo_ids}})
        cleanup_results["client_photos"] = photos_result.deleted_count
        print(f"   🗑️ Deleted {photos_result.deleted_count} orphaned client photos")
    else:
        cleanup_results["client_photos"] = 0
    
    total_deleted = sum(cleanup_results.values())
    print(f"🎯 Cleanup completed! Total deleted: {total_deleted} records")
    
    return {
        "message": "Orphaned data cleanup completed",
        "total_deleted": total_deleted,
        "details": cleanup_results,
        "remaining_trips": len(existing_trip_ids)
    }

# User Management - Block/Unblock and Delete functionality
@api_router.post("/users/{user_id}/block")
async def block_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin"]:
        raise HTTPException(status_code=403, detail="Only admins can block users")
    
    user_to_block = await db.users.find_one({"id": user_id})
    if not user_to_block:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user_to_block["role"] == "admin":
        raise HTTPException(status_code=403, detail="Cannot block admin users")
    
    await db.users.update_one({"id": user_id}, {"$set": {"blocked": True, "blocked_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "User blocked successfully"}

@api_router.post("/users/{user_id}/unblock")
async def unblock_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin"]:
        raise HTTPException(status_code=403, detail="Only admins can unblock users")
    
    await db.users.update_one({"id": user_id}, {"$set": {"blocked": False, "unblocked_at": datetime.now(timezone.utc).isoformat()}})
    return {"message": "User unblocked successfully"}

@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin"]:
        raise HTTPException(status_code=403, detail="Only admins can delete users")
    
    user_to_delete = await db.users.find_one({"id": user_id})
    if not user_to_delete:
        raise HTTPException(status_code=404, detail="User not found")
    
    if user_to_delete["role"] == "admin":
        raise HTTPException(status_code=403, detail="Cannot delete admin users")
    
    if user_to_delete["id"] == current_user["id"]:
        raise HTTPException(status_code=403, detail="Cannot delete yourself")
    
    # Delete user and related data
    await db.users.delete_one({"id": user_id})
    await db.client_notes.delete_many({"client_id": user_id})
    await db.quote_requests.delete_many({"client_id": user_id})
    
    return {"message": "User deleted successfully"}

# Notifications endpoint for payment deadlines
@api_router.get("/notifications/payment-deadlines")
async def get_payment_deadlines(current_user: dict = Depends(get_current_user)):
    if current_user["role"] not in ["admin", "agent"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    # Get upcoming payment deadlines (next 30 days)
    today = datetime.now(timezone.utc)
    thirty_days = today + timedelta(days=30)
    
    # Find payment installments due in next 30 days
    query = {
        "payment_date": {
            "$gte": today.isoformat(),
            "$lte": thirty_days.isoformat()
        }
    }
    
    # If agent, filter to their trips only
    if current_user["role"] == "agent":
        # Get agent's trip admin IDs
        agent_trips = await db.trips.find({"agent_id": current_user["id"]}).to_list(1000)
        trip_ids = [trip["id"] for trip in agent_trips]
        agent_trip_admins = await db.trip_admin.find({"trip_id": {"$in": trip_ids}}).to_list(1000)
        admin_ids = [admin["id"] for admin in agent_trip_admins]
        query["trip_admin_id"] = {"$in": admin_ids}
    
    upcoming_payments = await db.payment_installments.find(query).to_list(1000)
    
    # Get related trip and client information
    notifications = []
    for payment in upcoming_payments:
        # Get trip admin data
        trip_admin = await db.trip_admin.find_one({"id": payment["trip_admin_id"]})
        if not trip_admin:
            continue
        
        # Get trip data
        trip = await db.trips.find_one({"id": trip_admin["trip_id"]})
        if not trip:
            continue
        
        # Get client data
        client = await db.users.find_one({"id": trip["client_id"]})
        if not client:
            continue
        
        # Calculate days until due
        try:
            payment_date_str = payment["payment_date"]
            if isinstance(payment_date_str, str):
                payment_date = datetime.fromisoformat(payment_date_str.replace('Z', '+00:00'))
            else:
                payment_date = payment_date_str
                if payment_date.tzinfo is None:
                    payment_date = payment_date.replace(tzinfo=timezone.utc)
            
            days_until_due = (payment_date - today).days
        except Exception as e:
            print(f"Date parsing error for payment {payment.get('id', 'unknown')}: {e}")
            continue
        
        notifications.append({
            "id": payment["id"],
            "type": "payment_deadline",
            "title": f"Pagamento {payment['payment_type']} in scadenza",
            "message": f"Cliente {client['first_name']} {client['last_name']} - {trip['title']}",
            "amount": payment["amount"],
            "payment_date": payment["payment_date"],
            "days_until_due": days_until_due,
            "priority": "high" if days_until_due <= 7 else "medium" if days_until_due <= 14 else "low",
            "client_name": f"{client['first_name']} {client['last_name']}",
            "trip_title": trip["title"],
            "trip_id": trip["id"],
            "payment_type": payment["payment_type"]
        })
    
    # Sort by priority and days until due
    priority_order = {"high": 0, "medium": 1, "low": 2}
    notifications.sort(key=lambda x: (priority_order[x["priority"]], x["days_until_due"]))
    
    return {
        "notifications": notifications,
        "total_count": len(notifications),
        "high_priority_count": len([n for n in notifications if n["priority"] == "high"]),
        "medium_priority_count": len([n for n in notifications if n["priority"] == "medium"]),
        "low_priority_count": len([n for n in notifications if n["priority"] == "low"])
    }

# Include router
app.include_router(api_router)

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()